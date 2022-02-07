def teacher_hours(timetable_path, interestname):
  import openpyxl
  import re

  def stvis (string1): #функция приведения строк к общему виду
    str(string1)
    string2 = ''
    n = int(len(string1))
    for count in range(n):
      if (string1[count] != "\'") and (string1[count] != '[') and (string1[count] != ']') and (string1[count] != ' ') and (string1[count] != ','):
        string2 += str(string1[count])
    return string2

  book = openpyxl.load_workbook(timetable_path) #Открытие файла
  sheet = book['Лист1'] #Чтение первого листа

  days = [{}, {}, {}, {}, {}, {}, {}]
  week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]

  strpair = '' #создаю строку, куда все писать буду
  i = int(13) #ну, счетчик, чтоб был, дальше нужен. 13, ибо с этой строки считаем, нумерация строк/столбцов с 1 идет.
  timenow = sheet.cell(i, 19) #первая ячейка со временем пары
  timelast = sheet.cell(i + 1, 19) #Следующая ячейка из колонки со временем
  timeearl = timenow
  time = re.compile('^\s*\d{1,2}\.\d{2}\s*\-\s*\d{1,2}\.\d{2}\s*') #попытка задать регулярное выражение
  j = int(0) #счетчик для пар в день

  while (i < 133):
    pairnumber = int(1)
    while True:
      while True: #я хз, что оно делает, атцтаньти. Такая конструция, ибо цикла с пост-условием в этом языке вроде нет
        cell = sheet.cell(i, 20) #значение ячейки с парой/преподом/аудиторией/хз
        if(cell.value != None): #присоединение полученного значения к общей строке, если ячейка не пустая
          strpair += str(cell.value)
        i += 1 #Ну, очевидно
        if (timenow.value != None):
          timeearl = timenow
        timenow = sheet.cell(i, 19) #Обновляю ячейку времени
        enter = time.fullmatch(str(timenow.value)) #Ищу в взятой строке со временем вхождение регулярного выражения
        if(enter != None) or (i >= 132): #Если вхождение регулярного выражения в ячейку времени есть или случился конец таблицы, то ломаем цикл
          break
      days[j][pairnumber] = strpair #в словарь под ключ времени загоняем строку с парой
      pairnumber += 1
      if (i >= 132): #Проверка конца таблицы, ломаем, если да
        break
      strpair = '' #обнуляем строку пары
      timelast = sheet.cell(i, 19) #Смотрим следующее время
      flasttime = float(timelast.value[0:3]) #Переводим время начала следующей пары в число
      flearltime = float(timeearl.value[0:3]) #Переводим время начала предыдущей пары в число
      #print(flasttime, flearltime)
      #break
      if (flearltime > flasttime): #Сравниваем, если время начала предыдущей больше, то ломаем цикл
        break
    j += 1 #Переходим к следующему словарю внутри списка


  names = re.compile('[А-Я][а-я]{1,}\s*[А-Я]\.\s*[А-Я]\.') #Задаем регулярное выражение для ФИО преподавателя
  for j in range(6): #Ищем ФИО, если находим, то кидаем в массив
    for pairnumber in range(1, 6):
      name = names.findall(days[j][pairnumber])
      name.append(' ')
      name.append(' ')
      days[j][pairnumber] = name

  for j in range(6): #Разруливаем четность/нечетность недель, ну криво, ну что поделать
    for pairnumber in range(1, 6):
      if (days[j][pairnumber][1] == ' ') and (j != 3):
        days[j][pairnumber][1] = days[j][pairnumber][0]
      elif (j == 3) and (pairnumber == 4):
        days[j][pairnumber][1] = days[j][pairnumber][0]
        days[j][pairnumber][0] = ' '

  interestname = stvis(interestname)
  workhoursweek = [int(0), int(0)]
  result_hours = ''

  for j in range(6):
    workhoursday = [0, 0]
    for pairnumber in range(1, 6):
      for i in range(2):
        namestval = str(days[j][pairnumber][i])
        namestval = stvis(namestval)
        if(namestval == str(interestname)):
          workhoursday[i] += 1
    workhoursweek[0] += workhoursday[0]
    workhoursweek[1] += workhoursday[1]
    result_hours = result_hours + str(week[j]) + '\n' + 'Нечётная неделя ' + str(workhoursday[0]) + ' занятий/я | Чётная неделя ' + str(workhoursday[1]) + ' занятий/я' + '\n'
  result_hours += 'Всего занятий: \n'
  result_hours += 'Нечётная неделя ' + str(workhoursweek[0]) + ' занятий/я | Чётная неделя ' + str(workhoursweek[1]) + " занятий/я"
  return(result_hours)